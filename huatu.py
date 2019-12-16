import numpy as np
import matplotlib
import matplotlib.pyplot as plt#绘图
import os    
    
from openpyxl import load_workbook
import numpy

#加载Homework2.xlsx
wb = load_workbook('huitu.xlsx') 

#读取workbook中所有表格
sheets = wb.sheetnames
#打印所有表的名字
print(sheets)

#遍历每个sheet的数据
sheet1 = wb[sheets[0]]


#将sheet1的数据保存到num_list1
num_list1 = numpy.zeros((9,6))
i = 0
j = 0
for row in sheet1.rows:
    j=0
    for cell in row:
        num_list1[i][j] = cell.value
        j+=1
    i+=1
print(num_list1)
    

net_en_endata=num_list1[][0]
    
    
plt.title('Model loss',fontsize=20)
plt.ylabel('Loss',fontsize=20)
plt.xlabel('Epoch',fontsize=20)
plt.legend(loc='upper left')
plt.legend()  
plt.show()