# -*- coding: utf-8 -*-
"""
Created on Sat Mar  9 20:30:34 2019

@author: 99389
"""

from openpyxl import load_workbook
import numpy

#加载Homework2.xlsx
wb = load_workbook('excel.xlsx') 

#读取workbook中所有表格
sheets = wb.sheetnames
#打印所有表的名字
print(sheets)

#遍历每个sheet的数据
sheet1 = wb[sheets[0]]


#将sheet1的数据保存到num_list1
num_list1 = numpy.zeros((61,12))
i = 0
j = 0
for row in sheet1.rows:
    j=0
    for cell in row:
        num_list1[i][j] = cell.value
        j+=1
    i+=1
print(num_list1)
#coal
for i in range(61):
    num_list1[i][3]=num_list1[i][0]/(2000-num_list1[i][1]-num_list1[i][2]+num_list1[i][0])
    num_list1[i][4]=num_list1[i][0]/1000
    num_list1[i][5]=2*num_list1[i][3]*num_list1[i][4]/(num_list1[i][3]+num_list1[i][4])

#rock
for i in range (61):
    num_list1[i][6]=num_list1[i][1]/(2000-num_list1[i][0]-num_list1[i][2]+num_list1[i][1])
    num_list1[i][7]=num_list1[i][1]/(1000)
    num_list1[i][8]=2*num_list1[i][6]*num_list1[i][7]/(num_list1[i][6]+num_list1[i][7])

# coal_rock
for i in range (61):
    num_list1[i][9]=num_list1[i][2]/(2000-num_list1[i][0]-num_list1[i][1]+num_list1[i][2])
    num_list1[i][10]=num_list1[i][2]/1000
    num_list1[i][11]=2*num_list1[i][9]*num_list1[i][10]/(num_list1[i][9]+num_list1[i][10])


print(num_list1)

from xlwt import *
file = Workbook(encoding = 'utf-8')
#指定file以utf-8的格式打开
table = file.add_sheet('data')


for i,p in enumerate(num_list1):
#将数据写入文件,i是enumerate()函数返回的序号数
  for j,q in enumerate(p):
    # print i,j,q
    table.write(i,j,q)
file.save('data.xls')
